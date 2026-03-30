"""Generate a styled Excel workbook with native formulas."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from instapipe.metrics import Report


# ── Dark palette ─────────────────────────────────────────────────────────────
BG_DARK    = "FF0F0F1A"
ACCENT     = "FFB388EB"
ACCENT2    = "FFE6A8D7"
ROW1       = "FF1A1A2E"
ROW2       = "FF16213E"
GOLD       = "FFFCD34D"
WHITE      = "FFE2E8F0"

_fill = lambda h: PatternFill("solid", fgColor=h)
ACCENT_FILL   = _fill("FF2D1B69")
LAVENDER_FILL = _fill("FF3B0764")
ROW_FILL_1    = _fill(ROW1)
ROW_FILL_2    = _fill(ROW2)

title_font  = Font(name="Calibri", bold=True,  size=13, color=ACCENT)
sub_font    = Font(name="Calibri", bold=True,  size=11, color=ACCENT2)
head_font   = Font(name="Calibri", bold=True,  size=10, color=GOLD)
bold_font   = Font(name="Calibri", bold=True,  size=10, color=WHITE)
normal_font = Font(name="Calibri", bold=False, size=10, color=WHITE)

def _border():
    s = Side(style="thin", color="FF334155")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell):
    cell.font = head_font
    cell.fill = ACCENT_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def _cell(cell, bold=False, fill=None, fmt=None, align="right"):
    cell.font = bold_font if bold else normal_font
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt


def build_excel(report: Report, save_to: str | Path) -> None:
    """Generate a styled Excel workbook from a Report.

    Sheets:
        1. Overview — KPIs and top 5 reels
        2. Reels Raw — all reels with raw metrics
        3. Engagement Calc — native Excel formulas for engagement metrics
        4. Advanced Metrics — follower rate/1K, reach rate, quality score

    Args:
        report: Report object from metrics.compute().
        save_to: Path for the output .xlsx file.
    """
    df = report.data.copy()
    df["engagement_rate"] = report.engagement_rate
    df["save_rate"] = report.save_rate
    df["share_rate"] = report.share_rate
    df["follower_conv_rate"] = report.follower_conv_rate

    # Resolve column names
    desc_col = _resolve(df, ["descripcion_corta", "descripcion", "description"])
    views_col = _resolve(df, ["visualizaciones", "views"])
    reach_col = _resolve(df, ["alcance", "reach"])
    likes_col = _resolve(df, ["me_gustas", "likes"])
    comments_col = _resolve(df, ["comentarios", "comments"])
    saves_col = _resolve(df, ["guardados", "saves"])
    shares_col = _resolve(df, ["compartidos", "shares"])
    followers_col = _resolve(df, ["seguidores_ganados", "followers"])
    date_col = _resolve(df, ["fecha"])
    hour_col = _resolve(df, ["franja", "hora"])
    topic_col = _resolve(df, ["tema", "topic"])

    df = df.sort_values(views_col or "engagement_rate", ascending=False).reset_index(drop=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ═══ SHEET 1: OVERVIEW ═══════════════════════════════════════════════════
    ws1 = wb.create_sheet("Overview")
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:E1")
    ws1["A1"] = "Instagram Analytics — instapipe"
    ws1["A1"].font = title_font; ws1["A1"].fill = ACCENT_FILL
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35
    for w, c in zip([22, 18, 14, 14, 34], "ABCDE"):
        ws1.column_dimensions[c].width = w

    for i, h in enumerate(["Metric", "Value", "", "", "Notes"], 1):
        _hdr(ws1.cell(row=3, column=i, value=h))

    kpis = [
        ("Total reels", len(df)),
        ("Total views", int(df[views_col].sum()) if views_col else "N/A"),
        ("Total reach", int(df[reach_col].sum()) if reach_col else "N/A"),
        ("Total likes", int(df[likes_col].sum()) if likes_col else "N/A"),
        ("Total saves", int(df[saves_col].sum()) if saves_col else "N/A"),
        ("Total shares", int(df[shares_col].sum()) if shares_col else "N/A"),
        ("Followers gained", int(df[followers_col].sum()) if followers_col else "N/A"),
        ("Avg engagement rate", f"{report.engagement_rate.mean():.1f}%"),
        ("Avg save rate", f"{report.save_rate.mean():.2f}%"),
        ("Avg share rate", f"{report.share_rate.mean():.2f}%"),
    ]
    for r, (metric, val) in enumerate(kpis, 4):
        fill = ROW_FILL_1 if r % 2 == 0 else ROW_FILL_2
        ws1.cell(row=r, column=1, value=metric).font = bold_font
        ws1.cell(row=r, column=1).fill = fill; ws1.cell(row=r, column=1).border = _border()
        c = ws1.cell(row=r, column=2, value=val)
        _cell(c, bold=True, fill=fill, fmt="#,##0" if isinstance(val, int) else None)

    # ═══ SHEET 2: REELS RAW ══════════════════════════════════════════════════
    ws2 = wb.create_sheet("Reels Raw")
    ws2.sheet_view.showGridLines = False

    headers2 = ["#", "Date", "Topic", "Views", "Reach", "Likes", "Comments",
                "Saves", "Shares", "Followers", "ER %", "Save %", "Share %", "Description"]
    widths2 = [4, 12, 16, 12, 12, 10, 10, 10, 10, 10, 8, 8, 8, 40]
    for i, (h, w) in enumerate(zip(headers2, widths2), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        _hdr(ws2.cell(row=1, column=i, value=h))

    for r, (_, row) in enumerate(df.iterrows(), 2):
        fill = ROW_FILL_1 if r % 2 == 0 else ROW_FILL_2
        vals = [
            r - 1,
            str(row.get(date_col, "")) if date_col else "",
            str(row.get(topic_col, "")) if topic_col else "",
            int(row[views_col]) if views_col else 0,
            int(row[reach_col]) if reach_col else 0,
            int(row[likes_col]) if likes_col else 0,
            int(row[comments_col]) if comments_col else 0,
            int(row[saves_col]) if saves_col else 0,
            int(row[shares_col]) if shares_col else 0,
            int(row[followers_col]) if followers_col else 0,
            round(row["engagement_rate"], 2),
            round(row["save_rate"], 2),
            round(row["share_rate"], 2),
            str(row.get(desc_col, ""))[:50] if desc_col else "",
        ]
        for c, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = normal_font; cell.fill = fill; cell.border = _border()
            cell.alignment = Alignment(horizontal="left" if c in (2,3,14) else "right")
            if c in (4,5,6,7,8,9,10): cell.number_format = "#,##0"
            if c in (11,12,13): cell.number_format = "0.00"

    # ═══ SHEET 3: ENGAGEMENT CALC (FORMULAS) ═════════════════════════════════
    ws3 = wb.create_sheet("Engagement Calc")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:J1")
    ws3["A1"] = "Engagement Calculator — Native Excel Formulas"
    ws3["A1"].font = title_font; ws3["A1"].fill = ACCENT_FILL
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers3 = ["#", "Description", "Reach", "Likes", "Comments", "Saves",
                "Shares", "ER %", "Save Rate %", "Share Rate %"]
    widths3 = [4, 36, 12, 10, 10, 10, 12, 12, 12, 12]
    for i, (h, w) in enumerate(zip(headers3, widths3), 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        _hdr(ws3.cell(row=2, column=i, value=h))

    for r, (_, row) in enumerate(df.iterrows(), 3):
        fill = ROW_FILL_1 if r % 2 == 0 else ROW_FILL_2
        raw = [
            (1, r - 2),
            (2, str(row.get(desc_col, ""))[:40] if desc_col else ""),
            (3, int(row[reach_col]) if reach_col else 0),
            (4, int(row[likes_col]) if likes_col else 0),
            (5, int(row[comments_col]) if comments_col else 0),
            (6, int(row[saves_col]) if saves_col else 0),
            (7, int(row[shares_col]) if shares_col else 0),
        ]
        for c, val in raw:
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = normal_font; cell.fill = fill; cell.border = _border()
            cell.alignment = Alignment(horizontal="left" if c == 2 else "right")
            if c >= 3: cell.number_format = "#,##0"

        # ER = (likes + comments + saves + shares) / reach * 100
        c = ws3.cell(row=r, column=8)
        c.value = f"=IF(C{r}=0,0,(D{r}+E{r}+F{r}+G{r})/C{r}*100)"
        c.font = bold_font; c.fill = LAVENDER_FILL; c.border = _border()
        c.number_format = "0.00"

        # Save Rate = saves / reach * 100
        c = ws3.cell(row=r, column=9)
        c.value = f"=IF(C{r}=0,0,F{r}/C{r}*100)"
        c.font = normal_font; c.fill = fill; c.border = _border()
        c.number_format = "0.00"

        # Share Rate = shares / reach * 100
        c = ws3.cell(row=r, column=10)
        c.value = f"=IF(C{r}=0,0,G{r}/C{r}*100)"
        c.font = normal_font; c.fill = fill; c.border = _border()
        c.number_format = "0.00"

    # Summary row
    sr = len(df) + 3
    ws3.cell(row=sr, column=1, value="AVG").font = Font(bold=True, color=GOLD, size=11)
    ws3.cell(row=sr, column=1).fill = ACCENT_FILL; ws3.cell(row=sr, column=1).border = _border()
    for col in range(3, 11):
        ltr = get_column_letter(col)
        cell = ws3.cell(row=sr, column=col)
        cell.value = f"=SUM({ltr}3:{ltr}{sr-1})" if col <= 7 else f"=AVERAGE({ltr}3:{ltr}{sr-1})"
        cell.number_format = "#,##0" if col <= 7 else "0.00"
        cell.font = Font(bold=True, color=GOLD, size=10)
        cell.fill = ACCENT_FILL; cell.border = _border()

    # Save
    save_to = Path(save_to)
    save_to.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(save_to))


def _resolve(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Find the first column name that exists in the DataFrame."""
    for c in candidates:
        if c in df.columns:
            return c
    return None
